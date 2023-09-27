import requests
from bs4 import BeautifulSoup
import html5lib
import openpyxl
import pandas as pd

Headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36'}
data = requests.get('https://www.imdb.com/chart/top/', headers=Headers)
soup = BeautifulSoup(data.text, "html.parser")


excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank','Movie Name','Year','IMDB Rating'])


Movies = soup.find('ul',class_='ipc-metadata-list ipc-metadata-list--dividers-between sc-3f13560f-0 sTTRj compact-list-view ipc-metadata-list--base').find_all('li')


for Movie in Movies:
    names = Movie.find('div',class_='sc-4dcdad14-0 hqxhHZ cli-children').a.h3.get_text().split(' ')[1:4]
    merge = " ".join(names) 
    
    rank = Movie.find('div',class_='sc-4dcdad14-0 hqxhHZ cli-children').a.h3.get_text().split('.')[0]
    
    years = Movie.find('div',class_='sc-4dcdad14-7 enzKXX cli-title-metadata').span.get_text()
    
    ratings = Movie.find('span',class_='ipc-rating-star ipc-rating-star--base ipc-rating-star--imdb ratingGroup--imdb-rating').get_text()
    
    print(rank,merge,years,ratings)
    sheet.append([rank,merge,years,ratings])

excel.save('IMDB Movie Ratings.xlsx')




